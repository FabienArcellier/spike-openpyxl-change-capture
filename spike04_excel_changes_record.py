from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List, Tuple

@dataclass(frozen=True, slots=True)
class ExcelModel:
    sheet: str
    offset_row: int = 0
    offset_col: int = 0
    has_header: bool = True
    index: List[int] | List[str] | None = None # Optional index columns (integer if there is no header, string if there is header)

@dataclass(frozen=True, slots=True)
class ExcelRecord:
    sheet: str
    index: dict | int # Index data for the record (row if no index is specified in excel model or dictionary if index is specified)
    record: dict

@dataclass(frozen=True, slots=True)
class ExcelRecordChange:
    sheet: str
    index: dict | int # Index data for the record (row if no index is specified in excel model or dictionary if index is specified)
    change: str # 'added', 'removed', 'updated'
    old_value: dict | None = None
    new_value: dict | None = None

def excel_load_records(file_path: str, excel_structure: List[ExcelModel] | None = None, sheet: str | None = None) -> List[ExcelRecord]:
    workbook = load_workbook(file_path, data_only=True)
    records = []

    # Determine which sheets to process
    if excel_structure is not None:
        models_to_process = [model for model in excel_structure if sheet is None or model.sheet == sheet]
    else:
        models_to_process = [ExcelModel(sheet=s, offset_row=0, offset_col=0, has_header=True) for s in workbook.sheetnames if sheet is None or s == sheet]

    for model in models_to_process:
        if model.sheet not in workbook.sheetnames:
            continue  # Skip sheets not found in the workbook

        sheet_obj = workbook[model.sheet]
        headers = []
        start_row = model.offset_row + 1  # Convert to 1-based index for rows

        # Read headers if available
        if model.has_header:
            headers = [cell.value if cell.value is not None else f"Column_{i + 1}" for i, cell in enumerate(sheet_obj[start_row])]
            start_row += 1  # Data starts after the header row
        else:
            headers = [i + 1 for i in range(sheet_obj.max_column)]

        # Load each row as an ExcelRecord
        for row in sheet_obj.iter_rows(min_row=start_row, min_col=model.offset_col + 1, values_only=True):
            record_data = {headers[i]: (str(value) if value is not None else "") for i, value in enumerate(row)}

            # Determine the index based on the model's index configuration
            if model.index:
                if model.has_header:
                    index_data = {index: record_data.get(index, "") for index in model.index}
                else:
                    index_data = {headers[i]: record_data.get(headers[i], "") for i in model.index}
            else:
                index_data = start_row  # Use row number as index if no index specified

            # Append the ExcelRecord with the specified index and record data
            records.append(ExcelRecord(sheet=model.sheet, index=index_data, record=record_data))
            start_row += 1  # Increment for each row read

    return records


def excel_record_changes(origin_records: List[ExcelRecord], new_records: List[ExcelRecord]) -> Tuple[List[ExcelRecordChange], List[ExcelRecordChange], List[ExcelRecordChange]]:
    # Maps to store records by their (sheet, index) for easy lookup
    origin_map = {(record.sheet, frozenset(record.index.items())): record for record in origin_records}
    new_map = {(record.sheet, frozenset(record.index.items())): record for record in new_records}

    added_records = []
    removed_records = []
    updated_records = []

    # Detect removed and updated records
    for key, origin_record in origin_map.items():
        if key not in new_map:
            # Record is in origin but not in new, so it was removed
            removed_records.append(
                ExcelRecordChange(
                    sheet=origin_record.sheet,
                    index=origin_record.index,
                    change="removed",
                    old_value=origin_record.record,
                    new_value=None
                )
            )
        else:
            # Record exists in both; check if it was updated
            new_record = new_map[key]
            if origin_record.record != new_record.record:
                # The record has been updated
                updated_records.append(
                    ExcelRecordChange(
                        sheet=origin_record.sheet,
                        index=origin_record.index,
                        change="updated",
                        old_value=origin_record.record,
                        new_value=new_record.record
                    )
                )

    # Detect added records
    for key, new_record in new_map.items():
        if key not in origin_map:
            # Record is in new but not in origin, so it was added
            added_records.append(
                ExcelRecordChange(
                    sheet=new_record.sheet,
                    index=new_record.index,
                    change="added",
                    old_value=None,
                    new_value=new_record.record
                )
            )

    return added_records, removed_records, updated_records

excel_models = [
    ExcelModel(sheet="Clients", index=["ID Client"]),
    ExcelModel(sheet="Villes", index=["ID Ville"]),
    ExcelModel(sheet="Entreprises", index=["ID Entreprise"]),
]

origin_excel_records = excel_load_records('urban_planning-01.xlsx', excel_models)
new_excel_records = excel_load_records('urban_planning-02.xlsx', excel_models)

new_record, removed_record, updated_record = excel_record_changes(origin_excel_records, new_excel_records)

print("New Records:")
for record in new_record:
    print(f"* {record}")

print("Removed Records:")
for record in removed_record:
    print(f"* {record}")

print("Updated Records:")
for record in updated_record:
    print(f"* {record}")
