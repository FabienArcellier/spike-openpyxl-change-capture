from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List, Tuple

@dataclass(frozen=True, slots=True)
class ExcelModel:
    sheet: str
    offset_row: int = 0
    offset_col: int = 0
    has_header: bool = True
    index: List[int] | List[str] | None = None # Optional index columns (integer if there is no header, string if there is header)

@dataclass(frozen=True, slots=True)
class ExcelRecord:
    sheet: str
    index: dict | int # Index data for the record (row if no index is specified in excel model or dictionary if index is specified)
    record: dict


def excel_load_records(file_path: str, excel_structure: List[ExcelModel] | None = None, sheet: str | None = None) -> List[ExcelRecord]:
    workbook = load_workbook(file_path, data_only=True)
    records = []

    # Determine which sheets to process
    if excel_structure is not None:
        models_to_process = [model for model in excel_structure if sheet is None or model.sheet == sheet]
    else:
        models_to_process = [ExcelModel(sheet=s, offset_row=0, offset_col=0, has_header=True) for s in workbook.sheetnames if sheet is None or s == sheet]

    for model in models_to_process:
        if model.sheet not in workbook.sheetnames:
            continue  # Skip sheets not found in the workbook

        sheet_obj = workbook[model.sheet]
        headers = []
        start_row = model.offset_row + 1  # Convert to 1-based index for rows

        # Read headers if available
        if model.has_header:
            headers = [cell.value if cell.value is not None else f"Column_{i + 1}" for i, cell in enumerate(sheet_obj[start_row])]
            start_row += 1  # Data starts after the header row
        else:
            headers = [i + 1 for i in range(sheet_obj.max_column)]

        # Load each row as an ExcelRecord
        for row in sheet_obj.iter_rows(min_row=start_row, min_col=model.offset_col + 1, values_only=True):
            record_data = {headers[i]: (str(value) if value is not None else "") for i, value in enumerate(row)}

            # Determine the index based on the model's index configuration
            if model.index:
                if model.has_header:
                    index_data = {index: record_data.get(index, "") for index in model.index}
                else:
                    index_data = {headers[i]: record_data.get(headers[i], "") for i in model.index}
            else:
                index_data = start_row  # Use row number as index if no index specified

            # Append the ExcelRecord with the specified index and record data
            records.append(ExcelRecord(sheet=model.sheet, index=index_data, record=record_data))
            start_row += 1  # Increment for each row read

    return records

excel_models = [
    ExcelModel(sheet="Clients", index=["ID Client"]),
    ExcelModel(sheet="Villes", index=["ID Ville"]),
    ExcelModel(sheet="Entreprises", index=["ID Entreprise"]),
]

excel_records = excel_load_records('urban_planning-01.xlsx', excel_models)

for cell in excel_records[:10]:  # Print the first 10 cells as a sample
    print(cell)
