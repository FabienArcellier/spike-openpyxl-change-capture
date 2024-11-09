from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List


@dataclass(slots=True)
class ExcelCell:
    sheet: str
    row: int
    col: int
    value: str


def load_excel_cells(file_path: str) -> List[ExcelCell]:
    workbook = load_workbook(file_path, data_only=True)
    cells = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows(min_row=1, min_col=1):
            for cell in row:
                excel_cell = ExcelCell(
                    sheet=sheet_name,
                    row=cell.row,  # 1-based index
                    col=cell.column,  # 1-based index
                    value=str(cell.value) if cell.value is not None else ""  # Convert None to an empty string
                )
                cells.append(excel_cell)

    return cells


excel_cells = load_excel_cells('urban_planning-01.xlsx')

for cell in excel_cells[:10]:  # Print the first 10 cells as a sample
    print(cell)
