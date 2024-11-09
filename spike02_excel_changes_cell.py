from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List, Tuple


@dataclass(frozen=True, slots=True)
class ExcelCell:
    sheet: str
    row: int
    col: int
    value: str

@dataclass(frozen=True, slots=True)
class ExcelChange:
    sheet: str
    row: int
    col: int
    change: str # 'added', 'removed', 'updated'
    old_value: str
    new_value: str


def excel_load_cells(file_path: str) -> List[ExcelCell]:
    workbook = load_workbook(file_path, data_only=True)
    cells = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows(min_row=1, min_col=1):
            for cell in row:
                excel_cell = ExcelCell(
                    sheet=sheet_name,
                    row=cell.row,  # 1-based index
                    col=cell.column,  # 1-based index
                    value=str(cell.value) if cell.value is not None else ""  # Convert None to an empty string
                )
                cells.append(excel_cell)

    return cells

def excel_changes(origin_cells: List[ExcelCell], new_cells: List[ExcelCell]) -> Tuple[List[ExcelChange], List[ExcelChange], List[ExcelChange]]:
    # Create dictionaries for quick lookup by (sheet, row, col) keys
    origin_dict = {(cell.sheet, cell.row, cell.col): cell for cell in origin_cells}
    new_dict = {(cell.sheet, cell.row, cell.col): cell for cell in new_cells}

    added = []
    removed = []
    updated = []

    # Check for added and updated cells
    for key, new_cell in new_dict.items():
        if key not in origin_dict:
            # Cell is in new_cells but not in origin_cells
            added.append(ExcelChange(
                sheet=new_cell.sheet,
                row=new_cell.row,
                col=new_cell.col,
                change="added",
                old_value="",
                new_value=new_cell.value
            ))
        else:
            # Cell is in both, check if value has changed
            origin_cell = origin_dict[key]
            if origin_cell.value != new_cell.value:
                updated.append(ExcelChange(
                    sheet=new_cell.sheet,
                    row=new_cell.row,
                    col=new_cell.col,
                    change="updated",
                    old_value=origin_cell.value,
                    new_value=new_cell.value
                ))

    # Check for removed cells
    for key, origin_cell in origin_dict.items():
        if key not in new_dict:
            removed.append(ExcelChange(
                sheet=origin_cell.sheet,
                row=origin_cell.row,
                col=origin_cell.col,
                change="removed",
                old_value=origin_cell.value,
                new_value=""
            ))

    return added, removed, updated

origin_excel_cells = excel_load_cells('urban_planning-01.xlsx')
new_excel_cells = excel_load_cells('urban_planning-02.xlsx')

# Compare the two lists of Excel cells
new_cells, removed_cells, updated_cells = excel_changes(origin_excel_cells, new_excel_cells)

print("New Cells:")
for record in new_cells:
    print(f"* {record}")

print("Removed Cells:")
for record in removed_cells:
    print(f"* {record}")

print("Updated Cells:")
for record in updated_cells:
    print(f"* {record}")
